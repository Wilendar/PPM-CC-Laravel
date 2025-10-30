# -*- coding: utf-8 -*-
"""
FAST pattern analysis for ETAP_05d FAZA 2 - Sampled dataset (every 5th row)
"""
import openpyxl
from collections import defaultdict, Counter

# Open large Excel file
file_path = r"D:\OneDrive - MPP TRADE\Skrypty\PPM-CC-Laravel\References\Produkty_Przykład_Large.xlsx"
print('Loading Excel file (read-only mode)...')
wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
ws = wb.active

print('====================================================================')
print('FAST PATTERN ANALYSIS - Sampled Dataset (every 5th row)')
print('====================================================================')

# Get vehicle columns
vehicle_columns = {}
print('\n1. Loading vehicle columns...')
for col in range(16, 137):  # P-EF
    cell_value = ws.cell(1, col).value
    if cell_value:
        vehicle_columns[col] = cell_value

print(f'   Total vehicles: {len(vehicle_columns)}')

# Pattern stats
pattern_stats = {
    'both_types': 0,
    'only_original': 0,
    'only_replacement': 0,
    'no_compatibility': 0,
    'total_sampled': 0
}

compatibility_counts = []
vehicle_usage = Counter()
brand_families = defaultdict(set)
high_compat_products = []

print('\n2. Analyzing products (sampling every 5th row)...')
row = 2
sample_interval = 5

while row <= 500:  # Max 500 rows sampled (100 products analyzed)
    sku = ws.cell(row, 1).value

    if not sku:
        break

    pattern_stats['total_sampled'] += 1

    has_o = False
    has_z = False
    o_count = 0
    z_count = 0

    # Check vehicle columns
    for col in range(16, 137):
        val = ws.cell(row, col).value
        if val and str(val).upper() in ['O', 'Z']:
            vehicle_name = vehicle_columns.get(col, '')

            if str(val).upper() == 'O':
                has_o = True
                o_count += 1
            else:
                has_z = True
                z_count += 1

            vehicle_usage[vehicle_name] += 1

            # Brand family
            brand = vehicle_name.split()[0] if vehicle_name else ''
            if brand:
                brand_families[brand].add(vehicle_name)

    # Classify
    if has_o and has_z:
        pattern_stats['both_types'] += 1
    elif has_o:
        pattern_stats['only_original'] += 1
    elif has_z:
        pattern_stats['only_replacement'] += 1
    else:
        pattern_stats['no_compatibility'] += 1

    total_compat = o_count + z_count
    compatibility_counts.append(total_compat)

    if total_compat >= 20:
        high_compat_products.append((sku, total_compat, o_count, z_count))

    row += sample_interval

print(f'   Sampled products: {pattern_stats["total_sampled"]}')

# Results
print('\n====================================================================')
print('3. PATTERN STATISTICS:')
print('====================================================================')

total = pattern_stats["total_sampled"]
print(f'\nCompatibility Distribution:')
print(f'  Both (O + Z):   {pattern_stats["both_types"]:3d} ({pattern_stats["both_types"]/total*100:5.1f}%)')
print(f'  Only Original:  {pattern_stats["only_original"]:3d} ({pattern_stats["only_original"]/total*100:5.1f}%)')
print(f'  Only Zamiennik: {pattern_stats["only_replacement"]:3d} ({pattern_stats["only_replacement"]/total*100:5.1f}%)')
print(f'  No compat:      {pattern_stats["no_compatibility"]:3d} ({pattern_stats["no_compatibility"]/total*100:5.1f}%)')

if compatibility_counts:
    non_zero = [c for c in compatibility_counts if c > 0]
    print(f'\nCompatibility Counts:')
    print(f'  Average (all):      {sum(compatibility_counts)/len(compatibility_counts):5.1f} vehicles')
    if non_zero:
        print(f'  Average (non-zero): {sum(non_zero)/len(non_zero):5.1f} vehicles')
        print(f'  Median:             {sorted(non_zero)[len(non_zero)//2]:5d} vehicles')
        print(f'  Max:                {max(non_zero):5d} vehicles')
        print(f'  Min (non-zero):     {min(non_zero):5d} vehicles')

print('\n====================================================================')
print('4. TOP 15 MOST USED VEHICLES:')
print('====================================================================')
for vehicle, count in vehicle_usage.most_common(15):
    print(f'{vehicle:40s} | Uses: {count:3d}')

print('\n====================================================================')
print('5. BRAND FAMILIES:')
print('====================================================================')
for brand in sorted(brand_families.keys())[:20]:
    count = len(brand_families[brand])
    print(f'{brand:15s} | {count:3d} unique vehicles')

if high_compat_products:
    print('\n====================================================================')
    print(f'6. HIGH COMPATIBILITY PRODUCTS (>= 20 vehicles):')
    print('====================================================================')
    print(f'Found {len(high_compat_products)} products:\n')
    for sku, total, o, z in high_compat_products[:10]:
        print(f'SKU {sku:10s} | Total: {total:2d} | Oryginał: {o:2d} | Zamiennik: {z:2d}')

print('\n====================================================================')
print('7. UX INSIGHTS:')
print('====================================================================')

insights = []

# Mixed types
if pattern_stats['both_types'] > 0:
    pct = pattern_stats['both_types']/total*100
    insights.append(f'Mixed types: {pct:.1f}% products have BOTH O + Z')
    insights.append('  -> UX must support assigning different types to same product')

# High compat
if high_compat_products:
    insights.append(f'{len(high_compat_products)} products have >=20 compatibilities')
    insights.append('  -> Family helpers are CRITICAL')

# Brand families
if len(brand_families) > 10:
    insights.append(f'{len(brand_families)} brand families detected')
    insights.append('  -> Group vehicles by brand in search results')

# Avg compat
if non_zero:
    avg = sum(non_zero)/len(non_zero)
    if avg > 5:
        insights.append(f'Average {avg:.1f} vehicles per product')
        insights.append('  -> Bulk edit is ESSENTIAL (not optional)')

for i, insight in enumerate(insights, 1):
    print(f'{i}. {insight}')

print('\n====================================================================')
print('ANALYSIS COMPLETE')
print('====================================================================')

wb.close()
