# -*- coding: utf-8 -*-
"""
Deep pattern analysis for ETAP_05d FAZA 2 - Large Excel dataset
"""
import openpyxl
from collections import defaultdict, Counter
import sys

# Open large Excel file
file_path = r"D:\OneDrive - MPP TRADE\Skrypty\PPM-CC-Laravel\References\Produkty_Przykład_Large.xlsx"
wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
ws = wb.active

print('='*100)
print('DEEP PATTERN ANALYSIS - Large Dataset')
print('='*100)

# Get vehicle columns (P-EF = columns 16-136)
vehicle_columns = {}
print('\n1. VEHICLE COLUMNS MAPPING:')
for col in range(16, 137):
    cell_value = ws.cell(1, col).value
    if cell_value:
        col_letter = openpyxl.utils.get_column_letter(col)
        vehicle_columns[col] = cell_value

print(f'Total vehicles: {len(vehicle_columns)}')

# Analyze patterns
print('\n' + '='*100)
print('2. COMPATIBILITY PATTERNS (analyzing all products...)')
print('='*100)

# Pattern storage
pattern_stats = {
    'both_types': 0,  # Oryginał + Zamiennik
    'only_original': 0,
    'only_replacement': 0,
    'no_compatibility': 0,
    'total_products': 0
}

compatibility_counts = []
original_counts = []
replacement_counts = []

# Vehicle popularity
vehicle_usage = Counter()
vehicle_original = Counter()
vehicle_replacement = Counter()

# Brand families detection
brand_families = defaultdict(list)

# Product SKUs with high compatibility
high_compatibility_products = []  # (sku, count, types)

print('\nProcessing products...')
row = 2
processed = 0
while True:
    sku = ws.cell(row, 1).value  # Column A = SKU

    if not sku or row > 1000:  # Safety limit or end of data
        break

    pattern_stats['total_products'] += 1
    processed += 1

    if processed % 100 == 0:
        print(f'  Processed {processed} products...')

    has_original = False
    has_replacement = False
    original_count = 0
    replacement_count = 0
    product_vehicles = []

    # Check all vehicle columns
    for col in range(16, 137):
        cell_value = ws.cell(row, col).value

        if cell_value and str(cell_value).upper() in ['O', 'Z']:
            vehicle_name = vehicle_columns.get(col, '')

            if str(cell_value).upper() == 'O':
                has_original = True
                original_count += 1
                vehicle_original[vehicle_name] += 1
                product_vehicles.append((vehicle_name, 'O'))
            else:
                has_replacement = True
                replacement_count += 1
                vehicle_replacement[vehicle_name] += 1
                product_vehicles.append((vehicle_name, 'Z'))

            vehicle_usage[vehicle_name] += 1

            # Extract brand family
            brand = vehicle_name.split()[0] if vehicle_name else ''
            if brand:
                brand_families[brand].append(vehicle_name)

    # Classify pattern
    if has_original and has_replacement:
        pattern_stats['both_types'] += 1
    elif has_original:
        pattern_stats['only_original'] += 1
    elif has_replacement:
        pattern_stats['only_replacement'] += 1
    else:
        pattern_stats['no_compatibility'] += 1

    total_compat = original_count + replacement_count
    compatibility_counts.append(total_compat)
    original_counts.append(original_count)
    replacement_counts.append(replacement_count)

    # Track high compatibility products
    if total_compat >= 20:
        high_compatibility_products.append({
            'sku': sku,
            'total': total_compat,
            'original': original_count,
            'replacement': replacement_count,
            'vehicles': product_vehicles[:5]  # Sample
        })

    row += 1

print(f'\nTotal products analyzed: {pattern_stats["total_products"]}')

# Results
print('\n' + '='*100)
print('3. PATTERN STATISTICS:')
print('='*100)

print(f'\nCompatibility Distribution:')
print(f'  Both (Oryginał + Zamiennik): {pattern_stats["both_types"]} ({pattern_stats["both_types"]/pattern_stats["total_products"]*100:.1f}%)')
print(f'  Only Oryginał:               {pattern_stats["only_original"]} ({pattern_stats["only_original"]/pattern_stats["total_products"]*100:.1f}%)')
print(f'  Only Zamiennik:              {pattern_stats["only_replacement"]} ({pattern_stats["only_replacement"]/pattern_stats["total_products"]*100:.1f}%)')
print(f'  No compatibility:            {pattern_stats["no_compatibility"]} ({pattern_stats["no_compatibility"]/pattern_stats["total_products"]*100:.1f}%)')

print(f'\nCompatibility Counts per Product:')
if compatibility_counts:
    print(f'  Average: {sum(compatibility_counts)/len(compatibility_counts):.1f} vehicles')
    print(f'  Median: {sorted(compatibility_counts)[len(compatibility_counts)//2]} vehicles')
    print(f'  Max: {max(compatibility_counts)} vehicles')
    print(f'  Min (non-zero): {min([c for c in compatibility_counts if c > 0]) if any(c > 0 for c in compatibility_counts) else 0} vehicles')

print(f'\nOriginal vs Replacement Distribution:')
if original_counts and replacement_counts:
    print(f'  Avg Oryginał per product: {sum(original_counts)/len(original_counts):.1f}')
    print(f'  Avg Zamiennik per product: {sum(replacement_counts)/len(replacement_counts):.1f}')

print('\n' + '='*100)
print('4. TOP 10 MOST USED VEHICLES:')
print('='*100)
for vehicle, count in vehicle_usage.most_common(10):
    orig = vehicle_original[vehicle]
    repl = vehicle_replacement[vehicle]
    print(f'{vehicle:40s} | Total: {count:4d} | Oryginał: {orig:3d} | Zamiennik: {repl:3d}')

print('\n' + '='*100)
print('5. BRAND FAMILIES DETECTED:')
print('='*100)
for brand in sorted(brand_families.keys())[:15]:  # Top 15 brands
    unique_vehicles = set(brand_families[brand])
    print(f'{brand:15s} | {len(unique_vehicles):3d} unique vehicles')

print('\n' + '='*100)
print('6. HIGH COMPATIBILITY PRODUCTS (>=20 vehicles):')
print('='*100)
print(f'Found {len(high_compatibility_products)} products with high compatibility\n')
for prod in high_compatibility_products[:10]:  # Top 10
    print(f'SKU {prod["sku"]:10s} | Total: {prod["total"]:2d} | Oryginał: {prod["original"]:2d} | Zamiennik: {prod["replacement"]:2d}')
    print(f'  Sample vehicles:')
    for veh, typ in prod['vehicles']:
        print(f'    - {veh} ({typ})')

print('\n' + '='*100)
print('7. CRITICAL UX INSIGHTS:')
print('='*100)

insights = []

# Insight 1: Both types pattern
if pattern_stats['both_types'] > 0:
    pct = pattern_stats['both_types']/pattern_stats['total_products']*100
    insights.append(f'✓ {pct:.1f}% products have BOTH Oryginał + Zamiennik')
    insights.append('  → UX MUST support mixed types per product')

# Insight 2: High compatibility products
if high_compatibility_products:
    insights.append(f'✓ {len(high_compatibility_products)} products have >=20 compatibilities')
    insights.append('  → UX MUST have "Select all family" helpers')

# Insight 3: Brand families
if len(brand_families) > 10:
    insights.append(f'✓ {len(brand_families)} brand families detected')
    insights.append('  → UX MUST group vehicles by brand in search')

# Insight 4: Average compatibility
if compatibility_counts:
    avg = sum(compatibility_counts)/len(compatibility_counts)
    if avg > 5:
        insights.append(f'✓ Average {avg:.1f} vehicles per product')
        insights.append('  → Bulk operations are ESSENTIAL (not nice-to-have)')

for i, insight in enumerate(insights, 1):
    print(f'{i}. {insight}')

print('\n' + '='*100)
print('ANALYSIS COMPLETE')
print('='*100)

wb.close()
