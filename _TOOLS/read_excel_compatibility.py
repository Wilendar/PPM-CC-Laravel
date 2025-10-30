# -*- coding: utf-8 -*-
"""
Read Excel compatibility structure for PPM-CC-Laravel ETAP_05d FAZA 2 UX design
"""
import openpyxl
import sys

# Open Excel file
file_path = r"D:\OneDrive - MPP TRADE\Skrypty\PPM-CC-Laravel\References\Produkty_Przykład.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Print sheet name
print(f'Sheet: {ws.title}')
print('='*80)

# Get ALL column headers (row 1)
print('\nALL COLUMN HEADERS:')
all_headers = {}
for col in range(1, 137):  # A=1 to EF=136
    cell_value = ws.cell(1, col).value
    if cell_value:
        col_letter = openpyxl.utils.get_column_letter(col)
        all_headers[col_letter] = cell_value
        if col < 16:  # Before column P
            print(f'{col_letter}: {cell_value}')

print('\n' + '='*80)
print('\nVEHICLE COMPATIBILITY COLUMNS (P-EF):')
vehicle_columns = {}
for col in range(16, 137):  # P=16 to EF=136
    cell_value = ws.cell(1, col).value
    if cell_value:
        col_letter = openpyxl.utils.get_column_letter(col)
        vehicle_columns[col_letter] = cell_value
        print(f'{col_letter}: {cell_value}')

print(f'\nTotal vehicle columns: {len(vehicle_columns)}')

print('\n' + '='*80)
print('\nSAMPLE DATA (first 10 products):')
for row in range(2, 12):  # rows 2-11
    sku = ws.cell(row, 1).value  # column A = SKU
    name = ws.cell(row, 2).value  # column B = Name

    if not sku:
        continue

    print(f'\nRow {row}: {sku}')
    if name:
        print(f'  Name: {name[:50]}...' if len(str(name)) > 50 else f'  Name: {name}')

    # Check compatibility columns P-EF
    original_count = 0
    replacement_count = 0
    compatibilities = []

    for col in range(16, 137):  # P=16 to EF=136
        cell_value = ws.cell(row, col).value
        if cell_value in ['Z', 'O', 'z', 'o']:
            col_letter = openpyxl.utils.get_column_letter(col)
            vehicle = ws.cell(1, col).value

            comp_type = 'Oryginał' if cell_value.upper() == 'O' else 'Zamiennik'
            compatibilities.append(f'{vehicle}: {comp_type}')

            if cell_value.upper() == 'O':
                original_count += 1
            else:
                replacement_count += 1

    if compatibilities:
        print(f'  Oryginał: {original_count}, Zamiennik: {replacement_count}, Total: {len(compatibilities)} vehicles')
        print(f'  Sample compatibilities:')
        for comp in compatibilities[:5]:  # first 5
            print(f'    - {comp}')
        if len(compatibilities) > 5:
            print(f'    ... and {len(compatibilities) - 5} more')
    else:
        print('  No compatibilities')

print('\n' + '='*80)
print('\nWORKFLOW PATTERNS DETECTED:')
print('- Vertical drag: Multiple products (rows) → Same vehicle (column) → Same type (Z or O)')
print('- Horizontal drag: Single product (row) → Multiple vehicles (columns) → Same type (Z or O)')
print('- Mixed pattern: Product with both Z and O for different vehicles')

# Analyze patterns
print('\n' + '='*80)
print('\nPATTERN ANALYSIS (first 20 products):')
both_types = 0
only_original = 0
only_replacement = 0
no_compatibility = 0

for row in range(2, 22):
    sku = ws.cell(row, 1).value
    if not sku:
        continue

    has_original = False
    has_replacement = False

    for col in range(16, 137):
        cell_value = ws.cell(row, col).value
        if cell_value and cell_value.upper() == 'O':
            has_original = True
        elif cell_value and cell_value.upper() == 'Z':
            has_replacement = True

    if has_original and has_replacement:
        both_types += 1
    elif has_original:
        only_original += 1
    elif has_replacement:
        only_replacement += 1
    else:
        no_compatibility += 1

print(f'Both (Oryginał + Zamiennik): {both_types} products')
print(f'Only Oryginał: {only_original} products')
print(f'Only Zamiennik: {only_replacement} products')
print(f'No compatibility: {no_compatibility} products')

wb.close()
